import csv
import datetime
import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from builtins import input
import re


class Vacancy:
    currency_to_rub = {
        "KGS": 0.76,
        "AZN": 35.68,
        "BYR": 23.91,
        "EUR": 59.90,
        "GEL": 21.74,
        "KZT": 0.13,
        "RUR": 1,
        "UAH": 1.64,
        "USD": 60.66,
        "UZS": 0.0055,
    }

    def __init__(self, object_vacancy):
        self.name = object_vacancy['name']
        salary_from = (int)((float)("".join(object_vacancy['salary_from'].split())))
        salary_to = (int)((float)("".join(object_vacancy['salary_to'].split())))
        self.salary = (salary_from + salary_to) * self.currency_to_rub[object_vacancy['salary_currency']] // 2
        self.area_name = object_vacancy['area_name']
        self.published_at = datetime.datetime.strptime(object_vacancy['published_at'], '%Y-%m-%dT%H:%M:%S%z')


class DataSet:
    def __init__(self, file_name: str, vacancies_objects: list):
        self.file_name = file_name
        self.vacancies_objects = vacancies_objects

    def refactor_html(self, raw_html):
       return(re.compile('<.*?>'), '', raw_html)

    def csv_read(self):
        vacancies = []
        name_list = []
        with open(self.file_name, encoding='utf-8-sig') as r_file:
            file_reader = csv.reader(r_file, delimiter=",")
            count = 0
            for row in file_reader:
                if count == 0:
                    count += 1
                    name_list = row
                else:
                    if "" in row or len(row) != len(name_list):
                        continue
                    vacancies.append(row)
        if len(name_list) == 0:
            print('Пустой файл')
            exit()
        if len(vacancies) == 0:
            print('Нет данных')
            exit()
        return (vacancies, name_list)

    def fill_vacancies(self):
        (vacancies, list_naiming) = self.csv_read()
        self.vacancies_objects = self.csv_filer(vacancies, list_naiming)
        
    def csv_filer(self, reader, list_naming):
        vacancies = list()
        for row in reader:
            current = {}
            for i in range(len(row)):
                current[list_naming[i]] = row[i]
            vacancies.append(Vacancy(current))
        return vacancies



class Tuple:
    totalSalary = 0
    count = 0
    def __init__(self, totalSalary: int, count: int):
        self.totalSalary = totalSalary
        self.count = count


class InputData:
    years_stats = {
    }

    cities_stats = {
    }

    vacancy_stats = {
    }

    def start_input(self):
        self.file_name = input('Введите название файла: ')
        self.profession = input('Введите название профессии: ')
        self.city_count = 0

    def count_vacancies(self, vacancies: list):
        for vacancy in vacancies:
            self.city_count += 1
            year = int(vacancy.published_at.year)

            if vacancy.area_name not in self.cities_stats.keys():
                self.cities_stats[vacancy.area_name] = Tuple(vacancy.salary, 1)
            else:
                self.cities_stats[vacancy.area_name].totalSalary += vacancy.salary
                self.cities_stats[vacancy.area_name].count += 1
                
            if year not in self.years_stats.keys():
                self.years_stats[year] = Tuple(vacancy.salary, 1)
                self.vacancy_stats[year] = Tuple(0, 0)
            else:
                self.years_stats[year].totalSalary += vacancy.salary
                self.years_stats[year].count += 1

            if self.profession in vacancy.name:
                self.vacancy_stats[year].totalSalary += vacancy.salary
                self.vacancy_stats[year].count += 1

    def update_stats(self):
        for year in self.years_stats.keys():
            self.years_stats[year].totalSalary = int(self.years_stats[year].totalSalary // self.years_stats[year].count)

        remove_data = list()
        for city in self.cities_stats.keys():
            percentage = round(self.cities_stats[city].count / self.city_count, 4)
            if percentage < 0.01:
                remove_data.append(city)
            else:
                self.cities_stats[city].totalSalary = int(self.cities_stats[city].totalSalary // self.cities_stats[city].count)
                self.cities_stats[city].count = percentage
        for year in self.vacancy_stats.keys():
            if self.vacancy_stats[year].count != 0:
                self.vacancy_stats[year].totalSalary = int(self.vacancy_stats[year].totalSalary // self.vacancy_stats[year].count)
        for city in remove_data:
            del [self.cities_stats[city]]

    
    def print_info(self, str_info: str, dict: dict, value_name: str):
        marker = False
        print(str_info, end='')
        ind = 0
        for year in dict.keys():
            if ind == 0:
                print(' {', end='')
                ind += 1
            printEnd = ', '
            if year == max(dict.keys()):
                printEnd = ''
                marker = True
            print(f"{year}: {getattr(dict[year], value_name)}", end=printEnd)
        if marker:
            print('}')

    def get_city_print(self, str_data: str, dict: dict, names: list, value_name):
        flag = False
        print(str_data, end='')
        ind = 0
        for name in names:
            if ind == 0:
                print(' {', end='')
            printEnd = ', '
            if ind == len(names) - 1:
                printEnd = ''
                flag = True
            print(f"'{name}': {getattr(dict[name], value_name)}", end=printEnd)
            ind += 1
        if flag:
            print('}')

    def get_answer(self):

        cities_sorted = sorted(self.cities_stats, key=lambda x: self.cities_stats[x].totalSalary, reverse=True)
        self.print_info("Динамика уровня зарплат по годам:", self.years_stats, "totalSalary")
        self.print_info("Динамика уровня зарплат по годам для выбранной профессии:", self.vacancy_stats, "totalSalary")

        self.print_info("Динамика количества вакансий по годам:", self.years_stats, "count")

        self.print_info("Динамика количества вакансий по годам для выбранной профессии:", self.vacancy_stats, "count")
        del cities_sorted[10:]
        self.get_city_print("Уровень зарплат по городам (в порядке убывания):", self.cities_stats,
                              cities_sorted, "totalSalary")
        cities_sorted = sorted(self.cities_stats, key=lambda x: self.cities_stats[x].count, reverse=True)
        del cities_sorted[10:]
        self.get_city_print("Доля вакансий по городам (в порядке убывания):", self.cities_stats,
                              cities_sorted, "count")

    def get_sorted_cities(self, attr_name: str):
        current = {}
        sorted_names = sorted(self.cities_stats, key=lambda x: getattr(self.cities_stats[x], attr_name), reverse=True)
        del sorted_names[10:]
        for name in sorted_names:
            current[name] = self.cities_stats[name]
        return current


class Report:
    def generate_excel(self, inputer: InputData):
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        wb = openpyxl.Workbook()
        wb.remove(wb['Sheet'])
        wb.create_sheet("Статистика по годам")
        list = wb["Статистика по годам"]
        list['A1'].font = Font(bold=True)
        list['A1'].border = thin_border
        list['A1'] = "Год"
        list.column_dimensions['A'].width = 6
        list['B1'] = "Средняя зарплата"
        list['B1'].font = Font(bold=True)
        list['B1'].border = thin_border
        list.column_dimensions['B'].width = len("Средняя зарплата") + 2
        list['C1'] = f"Средняя зарплата - {inputer.profession}"
        list['C1'].font = Font(bold=True)
        list['C1'].border = thin_border
        list.column_dimensions['C'].width = len(f"Средняя зарплата - {inputer.profession}") + 2
        list['D1'] = "Количество вакансий"
        list['D1'].font = Font(bold=True)
        list['D1'].border = thin_border
        list.column_dimensions['D'].width = len("Количество вакансий") + 2
        list['E1'].font = Font(bold=True)
        list['E1'].border = thin_border
        list['E1'] = f"Количество вакансий - {inputer.profession}"
        list.column_dimensions['E'].width = len(f"Количество вакансий - {inputer.profession}") + 2
        for i in inputer.years_stats:
            list[f"A{i - 2005}"] = i
            list[f"A{i - 2005}"].border = thin_border
            list[f"B{i - 2005}"] = inputer.years_stats[i].totalSalary
            list[f"B{i - 2005}"].border = thin_border
            list[f"D{i - 2005}"] = inputer.years_stats[i].count
            list[f"D{i - 2005}"].border = thin_border
        for i in inputer.vacancy_stats:
            list[f"C{i - 2005}"] = inputer.vacancy_stats[i].totalSalary
            list[f"C{i - 2005}"].border = thin_border
            list[f"E{i - 2005}"] = inputer.vacancy_stats[i].count
            list[f"E{i - 2005}"].border = thin_border

        wb.create_sheet("Статистика по городам")
        list = wb["Статистика по городам"]
        list['A1'] = "Город"
        list['A1'].font = Font(bold=True)
        list['A1'].border = thin_border
        list.column_dimensions['A'].width = len("Город") + 2
        list['B1'] = "Уровень зарплат"
        list['B1'].font = Font(bold=True)
        list['B1'].border = thin_border
        list.column_dimensions['B'].width = len("Уровень зарплат") + 2
        list['D1'] = "Город"
        list['D1'].font = Font(bold=True)
        list['D1'].border = thin_border
        list.column_dimensions['D'].width = len("Город") + 2
        list['E1'] = "Доля ваканский"
        list['E1'].font = Font(bold=True)
        list['E1'].border = thin_border
        list.column_dimensions['E'].width = len("Доля ваканский") + 2
        sorted_cities = inputer.get_sorted_cities("totalSalary")
        ind = 2
        for i in sorted_cities:
            list[f"A{ind}"] = i
            list[f"A{ind}"].border = thin_border
            list.column_dimensions['A'].width = max(list.column_dimensions['A'].width, len(i) + 2)
            list[f"B{ind}"] = sorted_cities[i].totalSalary
            list[f"B{ind}"].border = thin_border
            ind = ind + 1
        sorted_cities = inputer.get_sorted_cities("count")
        ind = 2
        for i in sorted_cities:
            list[f"D{ind}"] = i
            list[f"D{ind}"].border = thin_border
            list.column_dimensions['D'].width = max(list.column_dimensions['D'].width, len(i) + 2)
            list[f"E{ind}"] = f"{round(sorted_cities[i].count * 100, 2)}%"
            list[f"E{ind}"].number_format = '0.00%'
            list[f"E{ind}"].border = thin_border
            ind = ind + 1
        wb.save("report.xlsx")





inputer = InputData()
inputer.start_input()
dataset = DataSet(inputer.file_name, list())
dataset.fill_vacancies()
inputer.count_vacancies(dataset.vacancies_objects)
inputer.update_stats()
inputer.get_answer()

report = Report()
report.generate_excel(inputer)

